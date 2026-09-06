import asyncio
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from playwright.async_api import async_playwright

async def extract_tn_farmer_schemes():
    excel_filename = "TN_Agri_GoScheme.xlsx"
    target_url = "https://www.tnagrisnet.tn.gov.in/people_app/GoScheme"

    # Initialize Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GoScheme Data"
    ws.views.sheetView[0].showGridLines = True

    # Exact headers matching the portal layout
    headers = [
        "S.No", 
        "Department", 
        "Scheme", 
        "Scheme Description", 
        "Eligibility Criteria", 
        "Document's Required", 
        "GO & Guidelines"
    ]
    ws.append(headers)

    # Style definitions
    header_font = Font(name="Arial", size=10, bold=True, color="000000")
    normal_font = Font(name="Arial", size=10, color="000000")
    blue_link_font = Font(name="Arial", size=10, color="0275D8", underline="single")
    
    thin_border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = thin_border

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={"width": 1400, "height": 900})
        page = await context.new_page()

        print(f"Navigating to {target_url}...")
        await page.goto(target_url, wait_until="networkidle", timeout=60000)

        # 1. Change entries per page dropdown to 50 or 100 (if present)
        try:
            page_size_select = await page.query_selector("select[name*='length'], select[name*='size'], .dataTables_length select")
            if page_size_select:
                print("Setting page size dropdown to 50 entries...")
                await page_size_select.select_option(value="50")
                await page.wait_for_load_state("networkidle")
                await page.wait_for_timeout(1000)
        except Exception as e:
            print(f"Page size selector not changed or not present: {e}")

        # 2. Wait for table rows to render
        await page.wait_for_selector("table tbody tr", timeout=15000)
        rows = await page.query_selector_all("table tbody tr")

        print(f"Found {len(rows)} rows in the table.")


        all_table_rows = []
        
        # 3. Read up to 50 rows
        for idx, row in enumerate(rows):
            if len(all_table_rows) >= 50:
                break

            cols = await row.query_selector_all("td")
            #print(f"Row {idx + 1}: Found {len(cols)} columns. : {[await col.inner_text() for col in cols]}")
            if len(cols) >= 7:
                s_no = (await cols[0].inner_text()).strip()
                dept = (await cols[1].inner_text()).strip()
                if len(s_no) > 0 and len(dept) == 0:
                    continue  # Skip empty rows
                scheme = (await cols[2].inner_text()).strip()
                desc = (await cols[3].inner_text()).strip()
                eligibility = (await cols[4].inner_text()).strip()
                documents = (await cols[5].inner_text()).strip()
                go_guidelines = (await cols[6].inner_text()).strip()

                if s_no or scheme:
                    all_table_rows.append([s_no, dept, scheme, desc, eligibility, documents, go_guidelines])

        print(f"Successfully extracted {len(all_table_rows)} rows without pagination.")

        # 4. Populate Excel sheet
        for row_idx, row_data in enumerate(all_table_rows, start=2):
            ws.append(row_data)
            
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                    cell.font = normal_font
                elif col_idx == 3:  # Scheme Title
                    cell.font = blue_link_font
                elif col_idx == 7 and "View" in str(cell.value):
                    cell.font = blue_link_font
                else:
                    cell.font = normal_font

        # Adjust column widths
        col_widths = {"A": 6, "B": 16, "C": 26, "D": 32, "E": 26, "F": 32, "G": 14}
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        ws.freeze_panes = "A2"
        
        wb.save(excel_filename)
        print(f"Saved 50 rows into Excel: '{excel_filename}'")
        await browser.close()

if __name__ == "__main__":
    asyncio.run(extract_tn_farmer_schemes())